"""Chapter 6: Interest Rate Futures - cheapest-to-deliver (CTD) bond finder for T-bond futures, with yield curve sensitivity analysis"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from ch04_interest_rates.yield_curve_bootstrap import fetch_treasury_yields, compute_zero_rates
from scipy.interpolate import CubicSpline
import numpy as np
import yfinance as yf
import matplotlib.pyplot as plt
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from dataclasses import dataclass
from openpyxl import load_workbook


TCF_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "TCF.xlsx")

# Change these to switch contract
SECTION_HEADER = "U.S. TREASURY BOND FUTURES CONTRACT"                        # ZB (30-year)
# SECTION_HEADER = "20-YEAR U.S. TREASURY BOND FUTURES CONTRACT"              # TWE (20-year)
# SECTION_HEADER = 'LONG-TERM "ULTRA" U.S. TREASURY BOND FUTURES CONTRACT'    # UB (ultra)
DELIVERY_DATE = date(2026, 6, 1)  # first day of delivery month
FACE = 100  # 100 for pricing convention, the actual face value of all bonds is $100k
TICKER = "ZBM26.CBT" # Edit the ticker to get the corresponding futures price


def fetch_futures_price(ticker):
    data = yf.Ticker(ticker).history(period="5d")
    return data["Close"].iloc[-1]

@dataclass
class DeliverableBond:
    coupon: float
    maturity: date
    conversion_factor: float
    cusip: str = ""
    price: float = None
    delivery_cost: float = None


def load_basket(section_header=SECTION_HEADER, delivery_date=DELIVERY_DATE, path=TCF_PATH):
    """Parse the CME TCF spreadsheet and return the deliverable basket.
    Download from: https://www.cmegroup.com/trading/interest-rates/treasury-conversion-factors.html"""

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Conversion Factors"]

    # Locate the contract section header
    section_row = None
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value and str(cell.value).strip() == section_header:
            section_row = cell.row
            break
    if section_row is None:
        raise ValueError(f"Could not find '{section_header}' in spreadsheet")

    # Find the delivery month column in the header row
    header_row = section_row + 5
    delivery_col = None
    for cell in ws[header_row]:
        if isinstance(cell.value, datetime) and cell.value.date() == delivery_date:
            delivery_col = cell.column
            break
    if delivery_col is None:
        raise ValueError(f"Delivery month {delivery_date} not found in spreadsheet columns")

    # Read bond data rows to append data for each deliverable bond
    basket = []
    for row in ws.iter_rows(min_row=section_row + 6, max_col=delivery_col):
        coupon_val = row[2].value
        if coupon_val is None:
            break

        cf = row[delivery_col - 1].value  # need to substract one because delivery_col comes from cell.column, which is Excel-native and not 0-indexed
        if not isinstance(cf, (int, float)):
            continue  # "-----" means bond is not eligible for this delivery month

        maturity_dt = row[4].value
        cusip = row[5].value

        basket.append(DeliverableBond(
            coupon=coupon_val / 100,
            maturity=maturity_dt.date() if isinstance(maturity_dt, datetime) else maturity_dt,
            conversion_factor=cf,
            cusip=str(cusip),
        ))

    wb.close()
    return basket


def accrued_interest(coupon, last_coupon_date, settlement_date, next_coupon_date):
    """Accrued interest (actual/actual day count convention for Treasuries)"""

    days_accrued = (settlement_date - last_coupon_date).days
    days_in_period = (next_coupon_date - last_coupon_date).days
    accrued = (coupon / 2) * (days_accrued / days_in_period)
    return accrued


def price_bond(coupon, maturity_date, face, forward_curve_fn, settlement_date):
    """Price bond from settlement date, using a forward zero curve"""

    coupon_dates = []
    cf_date = maturity_date
    while cf_date > settlement_date:
        coupon_dates.append(cf_date)
        cf_date -= relativedelta(months=6)
    coupon_dates.reverse()                  # Since we worked backwards from maturity

    next_coupon_date = coupon_dates[0]
    last_coupon_date = next_coupon_date - relativedelta(months=6)
    accrued = accrued_interest(coupon, last_coupon_date, settlement_date, next_coupon_date)

    cash_flows = [coupon / 2 * face] * len(coupon_dates)
    cash_flows[-1] += face

    dirty_price = 0
    for cash_flow, coupon_date in zip(cash_flows, coupon_dates):
        t = (coupon_date - settlement_date).days / 365.25
        dirty_price += cash_flow * np.exp(-forward_curve_fn(t) * t)

    clean_price = dirty_price - accrued

    return clean_price


def make_forward_curve(spot_fn, t0):
    """Derive forward zero rates starting at t0 from today's spot zero curve.
    Returns a function: t (years from t0) -> continuously compounded forward zero rate."""

    r_t0 = float(spot_fn(t0))
    def forward_fn(t):
        T = t0 + t
        return (float(spot_fn(T)) * T - r_t0 * t0) / t
    return forward_fn


def sort_bonds(basket, futures_price, forward_curve_fn):

    for bond in basket:
        bond.price = price_bond(bond.coupon, bond.maturity, FACE, forward_curve_fn, DELIVERY_DATE)
        bond.delivery_cost = bond.price - futures_price * bond.conversion_factor

    return sorted(basket, key=lambda b: b.delivery_cost)


def ctd_sensitivity(basket, spot_zero_curve_fn, t0):
    """Price all bonds across a range of parallel yield curve shifts. Returns shifts array
    and a (n_bonds x n_shifts) matrix of (price / CF) values."""

    shifts = np.arange(-0.02, 0.0225, 0.0025)
    price_cf = np.zeros((len(basket), len(shifts)))

    for j, shift in enumerate(shifts):
        shifted_spot = lambda t, s=shift: spot_zero_curve_fn(t) + s
        shifted_forward = make_forward_curve(shifted_spot, t0)
        for i, bond in enumerate(basket):
            p = price_bond(bond.coupon, bond.maturity, FACE, shifted_forward, DELIVERY_DATE)
            price_cf[i, j] = p / bond.conversion_factor

    return shifts, price_cf


def plot_ctd_sensitivity(basket, shifts, price_cf):
    """Plot price/CF for each bond across yield shifts, highlighting CTD bonds and switch points."""
    shifts_bp = shifts * 10000
    ctd_idx = np.argmin(price_cf, axis=0)           # index of CTD bond at each shift level
    ctd_bond_indices = set(ctd_idx.tolist())        # bonds that are CTD at any point

    _, ax = plt.subplots(figsize=(12, 7))

    for i in range(len(basket)):
        if i not in ctd_bond_indices:
            ax.plot(shifts_bp, price_cf[i], color="lightgrey", linewidth=0.8, zorder=1)

    colors = plt.cm.tab10(np.linspace(0, 1, len(ctd_bond_indices)))
    for color, i in zip(colors, sorted(ctd_bond_indices)):
        bond = basket[i]
        ax.plot(shifts_bp, price_cf[i], color=color, linewidth=2,
                label=f"{bond.coupon*100:.3f}% {bond.maturity}", zorder=3)

    envelope = price_cf[ctd_idx, np.arange(len(shifts))]
    ax.plot(shifts_bp, envelope, color="black", linewidth=1.5, linestyle="--",
            label="CTD envelope", zorder=4)

    for sp in np.where(np.diff(ctd_idx) != 0)[0]:
        ax.axvline(x=shifts_bp[sp + 1], color="red", linestyle=":", linewidth=1, alpha=0.7, zorder=2)

    ax.axvline(x=0, color="black", linewidth=1, alpha=0.3, zorder=2)
    ax.set_xlabel("Yield shift (bps)")
    ax.set_ylabel("Price / Conversion factor")
    ax.set_title(f"CTD sensitivity - {TICKER}")
    ax.legend(fontsize=8, loc="upper right")
    ax.grid(True, linestyle="--", alpha=0.4)
    plt.tight_layout()
    plt.savefig(os.path.join(os.path.dirname(os.path.abspath(__file__)), "ctd_sensitivity.png"), dpi=150, bbox_inches="tight")
    plt.show()


if __name__ == "__main__":

    basket = load_basket()
    maturities, par_yields = fetch_treasury_yields()
    zero_rates = compute_zero_rates(maturities, par_yields)
    spot_zero_curve_fn = CubicSpline(maturities, zero_rates)
    futures_price = fetch_futures_price(TICKER)

    t0 = (DELIVERY_DATE - date.today()).days / 365.25
    forward_curve_fn = make_forward_curve(spot_zero_curve_fn, t0)

    sorted_basket = sort_bonds(basket, futures_price, forward_curve_fn)
    ctd = sorted_basket[0]
    print(f"CTD bond: {ctd.coupon*100:.3f}% {ctd.maturity} | Price: {ctd.price:.4f} | CF: {ctd.conversion_factor:.4f} | Delivery cost: {ctd.delivery_cost:.4f}\n")
    print(f"{'Coupon':<8}  {'Maturity':<10}  {'Price':>8}  {'CF':>8}  {'Del. cost':>10}  {'CUSIP'}")
    print("-" * 61)
    for bond in sorted_basket:
        marker = " <-- CTD" if bond is ctd else ""
        cp = f"{bond.coupon*100:.3f}%"
        print(f"{cp:<8}  {bond.maturity}  {bond.price:>8.4f}  {bond.conversion_factor:>8.4f}  {bond.delivery_cost:>10.4f}  {bond.cusip}{marker}")

    shifts, price_cf = ctd_sensitivity(basket, spot_zero_curve_fn, t0)

    ctd_idx = np.argmin(price_cf, axis=0)
    print("\nCTD by yield shift:")
    print(f"\n{'Shift':<8}  {'Coupon':<8}  {'Maturity':<10}  {'CUSIP'}")
    print("-" * 42)
    prev_idx = None
    for j, shift in enumerate(shifts):
        bond = basket[ctd_idx[j]]
        switch = "  <-- switch" if prev_idx is not None and ctd_idx[j] != prev_idx else ""
        cp = f"{bond.coupon*100:.3f}%"
        sh = f"{shift*10000:+.0f}bp"
        print(f"{sh:<8}  {cp:<8}  {bond.maturity}  {bond.cusip}{switch}")
        prev_idx = ctd_idx[j]

    plot_ctd_sensitivity(basket, shifts, price_cf)